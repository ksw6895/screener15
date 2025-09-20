import os
import asyncio
import aiohttp
import json
import time
import numpy as np
import statistics
from datetime import datetime
from aiohttp import ClientSession, ClientResponseError
import logging
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Constants for Rate Limiting
REQUEST_DELAY = 0.06  # Seconds between API requests
MAX_CONCURRENT_REQUESTS = 10

# Load configuration
CONFIG_FILE = 'enhanced_config.json'

def load_config(config_file):
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file '{config_file}' not found.")
    with open(config_file, 'r') as f:
        return json.load(f)

config = load_config(CONFIG_FILE)

# Setup logging
logging.basicConfig(
    level=getattr(logging, config['logging']['level'].upper(), logging.INFO),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(config['logging']['file']),
        logging.StreamHandler(sys.stdout)
    ]
)

API_KEY = config['api_key']
BASE_URL_V3 = config['base_url']
BASE_URL_V4 = "https://financialmodelingprep.com/api/v4"

if not API_KEY:
    logging.error("API_KEY not found in configuration.")
    raise ValueError("API_KEY not found in configuration.")

@dataclass
class FinancialMetrics:
    revenue: List[float]
    eps: List[float]
    fcf: List[float]
    ttm_fcf: float
    roe: List[float]
    gross_margin: List[float]
    operating_margin: List[float]
    working_capital: List[float]
    rd_expense: List[float]
    capex: List[float]
    per: List[float]
    pbr: List[float]
    dates: List[str]

class MetricNormalizer:
    def __init__(self, winsorize_percentile: float = 0.05):
        self.winsorize_percentile = winsorize_percentile
        self.metric_ranges = {}

    def winsorize(self, values: List[float]) -> List[float]:
        if not values:
            return values
        sorted_values = sorted(values)
        n = len(sorted_values)
        lower_idx = int(n * self.winsorize_percentile)
        upper_idx = int(n * (1 - self.winsorize_percentile))
        lower_bound = sorted_values[lower_idx]
        upper_bound = sorted_values[upper_idx]
        return [min(max(x, lower_bound), upper_bound) for x in values]

    def normalize(self, values: List[float], metric_name: str) -> List[float]:
        if not values:
            return values
        winsorized = self.winsorize(values)
        min_val = min(winsorized)
        max_val = max(winsorized)
        if max_val == min_val:
            return [1.0 if v >= max_val else 0.0 for v in values]
        self.metric_ranges[metric_name] = (min_val, max_val)
        return [(v - min_val) / (max_val - min_val) for v in values]

class GrowthQualityAnalyzer:
    def __init__(self, normalizer: MetricNormalizer, target_rates: Dict[str, float]):
        self.normalizer = normalizer
        self.target_rates = target_rates
        self.weights = {
            'magnitude': 0.35,
            'consistency': 0.35,
            'sustainability': 0.30
        }

    def calculate_growth_consistency(self, values: List[float]) -> float:
        if len(values) < 2:
            return 0.0
        growth_rates = [
            (values[i] - values[i - 1]) / abs(values[i - 1])
            for i in range(1, len(values))
            if values[i - 1] != 0
        ]
        if not growth_rates:
            return 0.0
        mean_growth = statistics.mean(growth_rates)
        if mean_growth == 0:
            return 0.0
        try:
            std_dev = statistics.stdev(growth_rates)
            cv = std_dev / abs(mean_growth)
            consistency = 1 / (1 + cv)
            if mean_growth < 0:
                consistency *= -1
            return consistency
        except statistics.StatisticsError:
            return 0.0

    def calculate_magnitude_score(self, actual_cagr: float, target_cagr: float) -> float:
        if target_cagr == 0:
            return 0.0
        score = (actual_cagr - target_cagr) / abs(target_cagr)
        score = max(-1.0, min(1.0, score))
        return score

    def calculate_trend_score(self, values: List[float]) -> float:
        if len(values) < 2:
            return 0.0
        changes = [
            (values[i] - values[i - 1]) / abs(values[i - 1]) if values[i - 1] != 0 else 0
            for i in range(1, len(values))
        ]
        if not changes:
            return 0.0
        consistency = self.calculate_growth_consistency(values)
        avg_change = sum(changes) / len(changes)
        trend_strength = 1 / (1 + np.exp(-avg_change))
        return (consistency + trend_strength) / 2

    def assess_growth_sustainability(self, metrics: FinancialMetrics) -> float:
        rd_intensity = [
            rd / rev if rev > 0 else 0
            for rd, rev in zip(metrics.rd_expense, metrics.revenue)
        ]
        rd_trend = self.calculate_trend_score(rd_intensity)

        capex_efficiency = [
            fcf / cap if cap > 0 else 0
            for fcf, cap in zip(metrics.fcf, metrics.capex)
        ]
        capex_trend = self.calculate_trend_score(capex_efficiency)

        margin_stability = self.calculate_growth_consistency(metrics.operating_margin)

        fcf_conversion = [
            fcf / rev if rev > 0 else 0
            for fcf, rev in zip(metrics.fcf, metrics.revenue)
        ]
        fcf_trend = self.calculate_trend_score(fcf_conversion)

        weights = {
            'rd_trend': 0.25,
            'capex_trend': 0.25,
            'margin_stability': 0.25,
            'fcf_trend': 0.25
        }

        sustainability_score = (
            weights['rd_trend'] * rd_trend +
            weights['capex_trend'] * capex_trend +
            weights['margin_stability'] * margin_stability +
            weights['fcf_trend'] * fcf_trend
        )

        return sustainability_score

    def normalize_growth_rate(self, growth_rate: float, sector_median: float, sector_std: float) -> float:
        if sector_std == 0:
            return 0.0
        return (growth_rate - sector_median) / sector_std

    def calculate_growth_scores(self, metrics: FinancialMetrics, sector_revenue_median: float = 0.1, sector_revenue_std: float = 0.05) -> Dict[str, float]:
        def calculate_cagr(end_value: float, start_value: float, periods: int) -> float:
            try:
                if start_value <= 0 or end_value <= 0 or periods <= 0:
                    return 0.0
                return (end_value / start_value) ** (1 / periods) - 1
            except (ValueError, ZeroDivisionError):
                return 0.0

        revenue_cagr = calculate_cagr(metrics.revenue[-1], metrics.revenue[0], len(metrics.revenue))
        # Normalize revenue CAGR to sector
        revenue_cagr_normalized = self.normalize_growth_rate(revenue_cagr, sector_revenue_median, sector_revenue_std)
        eps_cagr = calculate_cagr(metrics.eps[-1], metrics.eps[0], len(metrics.eps))
        fcf_cagr = calculate_cagr(metrics.fcf[-1], metrics.fcf[0], len(metrics.fcf))

        magnitude_scores = {
            'revenue': self.calculate_magnitude_score(revenue_cagr_normalized, self.target_rates.get('revenue', 0.1)),
            'eps': self.calculate_magnitude_score(eps_cagr, self.target_rates.get('eps', 0.1)),
            'fcf': self.calculate_magnitude_score(fcf_cagr, self.target_rates.get('fcf', 0.1))
        }

        magnitude_score = sum(magnitude_scores.values()) / len(magnitude_scores)

        consistency_scores = {
            'revenue': self.calculate_growth_consistency(metrics.revenue),
            'eps': self.calculate_growth_consistency(metrics.eps),
            'fcf': self.calculate_growth_consistency(metrics.fcf)
        }
        consistency_score = sum(consistency_scores.values()) / len(consistency_scores)

        sustainability_score = self.assess_growth_sustainability(metrics)

        final_score = (
            self.weights['magnitude'] * magnitude_score +
            self.weights['consistency'] * consistency_score +
            self.weights['sustainability'] * sustainability_score
        )

        return {
            'final_score': final_score,
            'magnitude_score': magnitude_score,
            'consistency_score': consistency_score,
            'sustainability_score': sustainability_score,
            'individual_magnitude_scores': magnitude_scores,
            'individual_consistency_scores': consistency_scores
        }

class RiskAssessmentModule:
    def __init__(self, normalizer: MetricNormalizer):
        self.normalizer = normalizer

    def calculate_stability_score(self, values: List[float]) -> float:
        if len(values) < 2:
            return 0.0
        try:
            mean = statistics.mean(values)
            if mean == 0:
                return 0.0
            std = statistics.stdev(values)
            cv = std / abs(mean)
            stability = 1 / (1 + cv)
            return stability
        except statistics.StatisticsError:
            return 0.0

    def calculate_trend_score(self, values: List[float]) -> float:
        if len(values) < 2:
            return 0.0
        changes = [
            values[i] - values[i - 1]
            for i in range(1, len(values))
        ]
        avg_change = sum(changes) / len(changes)
        trend_strength = 1 / (1 + np.exp(-avg_change))
        stability = self.calculate_stability_score(values)
        return (trend_strength + stability) / 2

    def calculate_working_capital_efficiency(self, metrics: FinancialMetrics) -> float:
        if not metrics.revenue or not metrics.working_capital:
            return 0.0
        wc_turnover = [
            rev / wc if wc > 0 else 0
            for rev, wc in zip(metrics.revenue, metrics.working_capital)
        ]
        normalized_turnover = self.normalizer.normalize(wc_turnover, 'wc_turnover')
        if not normalized_turnover:
            return 0.0
        trend_score = self.calculate_trend_score(normalized_turnover)
        return trend_score

    def calculate_margin_risk(self, metrics: FinancialMetrics) -> float:
        if not metrics.gross_margin or not metrics.operating_margin:
            return 0.0
        gross_margin_stability = self.calculate_stability_score(metrics.gross_margin)
        operating_margin_stability = self.calculate_stability_score(metrics.operating_margin)
        gross_margin_trend = self.calculate_trend_score(metrics.gross_margin)
        operating_margin_trend = self.calculate_trend_score(metrics.operating_margin)

        weights = {
            'gross_stability': 0.25,
            'operating_stability': 0.25,
            'gross_trend': 0.25,
            'operating_trend': 0.25
        }

        margin_risk_score = (
            weights['gross_stability'] * gross_margin_stability +
            weights['operating_stability'] * operating_margin_stability +
            weights['gross_trend'] * gross_margin_trend +
            weights['operating_trend'] * operating_margin_trend
        )

        return margin_risk_score

class ValuationAnalyzer:
    def __init__(self, normalizer: MetricNormalizer, config: dict):
        self.normalizer = normalizer
        self.config = config

    def calculate_valuation_score(self, per: float, pbr: float) -> float:
        per_score = self.normalize_per(per)
        pbr_score = self.normalize_pbr(pbr)
        valuation_score = (per_score + pbr_score) / 2
        return valuation_score

    def normalize_per(self, per: float) -> float:
        min_per = self.config.get('per_min', 5)
        max_per = self.config.get('per_max', 30)
        if per <= 0:
            return 0.0
        elif per < min_per:
            return 1.0
        elif per > max_per:
            return 0.0
        else:
            return 1 - ((per - min_per) / (max_per - min_per))

    def normalize_pbr(self, pbr: float) -> float:
        min_pbr = self.config.get('pbr_min', 0.5)
        max_pbr = self.config.get('pbr_max', 3)
        if pbr <= 0:
            return 0.0
        elif pbr < min_pbr:
            return 1.0
        elif pbr > max_pbr:
            return 0.0
        else:
            return 1 - ((pbr - min_pbr) / (max_pbr - min_pbr))

    def calculate_fcf_yield(self, ttm_fcf: float, market_cap: float) -> float:
        if market_cap <= 0:
            return 0.0
        fcf_yield = ttm_fcf / market_cap
        target_fcf_yield = self.config.get('fcf_yield_target', 0.05)
        score = fcf_yield / target_fcf_yield
        score = max(-1.0, min(1.0, score))
        return score

    def calculate_growth_adjusted_valuation(self, per: float, growth_rate: float) -> float:
        if growth_rate <= 0:
            return 0.0
        peg_ratio = per / (growth_rate * 100)  # growth_rate in decimal
        if peg_ratio > 5:
            return 0.0
        elif peg_ratio < 0.5:
            return 1.0
        else:
            return 1 - ((peg_ratio - 0.5) / (5 - 0.5))

class QualityScorer:
    def __init__(
        self, normalizer: MetricNormalizer,
        growth_analyzer: GrowthQualityAnalyzer,
        risk_assessor: RiskAssessmentModule,
        valuation_analyzer: ValuationAnalyzer
    ):
        self.normalizer = normalizer
        self.growth_analyzer = growth_analyzer
        self.risk_assessor = risk_assessor
        self.valuation_analyzer = valuation_analyzer

    def calculate_market_strength(self, metrics: FinancialMetrics) -> float:
        return 0.5

    def calculate_coherence_bonus(self, metrics: FinancialMetrics) -> float:
        growth_scores = self.growth_analyzer.calculate_growth_scores(metrics)
        growth_alignment = growth_scores['final_score']

        margin_risk = self.risk_assessor.calculate_margin_risk(metrics)
        wc_efficiency = self.risk_assessor.calculate_working_capital_efficiency(metrics)
        sustainability = growth_scores['sustainability_score']

        coherence_score = statistics.mean([
            growth_alignment,
            margin_risk,
            wc_efficiency,
            sustainability
        ])

        max_bonus = 0.10
        bonus_multiplier = 1 + (max_bonus * coherence_score)
        return bonus_multiplier

    def calculate_final_score(
        self, metrics: FinancialMetrics, market_cap: float, sector: str = "Technology"
    ) -> Tuple[float, Dict[str, float]]:
        component_scores = {}

        base_weights = {
            'growth': 0.60,
            'valuation_score': 0.10,
            'margin_quality': 0.10,
            'working_capital_efficiency': 0.10,
            'free_cash_flow_yield': 0.10
        }

        if sector == "Technology":
            base_weights['growth'] = 0.65
            base_weights['valuation_score'] = 0.05

        growth_scores = self.growth_analyzer.calculate_growth_scores(metrics)
        component_scores.update(growth_scores)

        per = metrics.per[0]
        pbr = metrics.pbr[0]
        valuation_score = self.valuation_analyzer.calculate_valuation_score(per, pbr)
        component_scores['valuation_score'] = valuation_score

        margin_quality = self.risk_assessor.calculate_margin_risk(metrics)
        wc_efficiency = self.risk_assessor.calculate_working_capital_efficiency(metrics)
        component_scores['margin_quality'] = margin_quality
        component_scores['working_capital_efficiency'] = wc_efficiency

        ttm_fcf = metrics.ttm_fcf
        fcf_yield_score = self.valuation_analyzer.calculate_fcf_yield(ttm_fcf, market_cap)
        component_scores['free_cash_flow_yield'] = fcf_yield_score

        base_score = (
            base_weights['growth'] * growth_scores['final_score'] +
            base_weights['valuation_score'] * valuation_score +
            base_weights['margin_quality'] * margin_quality +
            base_weights['working_capital_efficiency'] * wc_efficiency +
            base_weights['free_cash_flow_yield'] * fcf_yield_score
        )

        coherence_multiplier = self.calculate_coherence_bonus(metrics)
        final_score = base_score * coherence_multiplier

        component_scores['coherence_multiplier'] = coherence_multiplier
        component_scores['final_score'] = final_score

        return final_score, component_scores

def get_timestamp():
    return datetime.now().strftime('%Y%m%d_%H%M%S')

def calculate_cagr(end_value: float, start_value: float, periods: int) -> float:
    try:
        if start_value <= 0 or end_value <= 0 or periods <= 0:
            return 0.0
        return (end_value / start_value) ** (1 / periods) - 1
    except (ValueError, ZeroDivisionError):
        return 0.0

async def fetch(session, url, semaphore):
    wait_time = REQUEST_DELAY
    while True:
        try:
            async with semaphore:
                await asyncio.sleep(REQUEST_DELAY)
                async with session.get(url, timeout=10) as response:
                    if response.status == 200:
                        return await response.json()
                    elif response.status == 429:
                        logging.warning(f"Rate limited. Waiting {wait_time}s before retry: {url}")
                        await asyncio.sleep(wait_time)
                        wait_time = min(wait_time * 2, 60)
                    elif response.status == 404:
                        logging.error(f"Resource not found (404): {url}")
                        return None
                    else:
                        response.raise_for_status()
        except Exception as e:
            logging.error(f"Error fetching {url}: {str(e)}")
            await asyncio.sleep(wait_time)
            wait_time = min(wait_time * 2, 60)

def prepare_financial_metrics(
    income_statements: List[dict],
    cash_flow_statements: List[dict],
    ratios: List[dict],
    ratios_ttm: List[dict],
    key_metrics: List[dict],
    financial_growth: List[dict],
) -> Optional[FinancialMetrics]:
    if not income_statements or not cash_flow_statements or not ratios or not key_metrics:
        return None
    try:
        income_dict = {stmt['date']: stmt for stmt in income_statements}
        cash_flow_dict = {stmt['date']: stmt for stmt in cash_flow_statements}
        ratios_dict = {ratio['date']: ratio for ratio in ratios}
        key_metrics_dict = {metric['date']: metric for metric in key_metrics}

        # Extract TTM-based metrics once
        ttm_data = {}
        if ratios_ttm and isinstance(ratios_ttm, list) and len(ratios_ttm) > 0:
            ttm_data = ratios_ttm[0]
        # Current PER from TTM if available
        per_value_ttm = ttm_data.get('priceEarningsRatioTTM', ttm_data.get('peRatioTTM', 0))
        # Current PBR from TTM if available
        pbr_value_ttm = ttm_data.get('priceToBookRatioTTM', 0)

        common_dates = set(income_dict.keys()) & set(cash_flow_dict.keys()) & set(ratios_dict.keys()) & set(key_metrics_dict.keys())
        sorted_dates = sorted(common_dates, reverse=True)

        if not sorted_dates:
            logging.debug("No common dates between financial statements.")
            return None

        revenue = []
        eps = []
        fcf = []
        roe = []
        gross_margin = []
        operating_margin = []
        working_capital = []
        rd_expense = []
        capex = []
        per = []
        pbr = []
        dates = []

        for date in sorted_dates:
            income_stmt = income_dict.get(date, {})
            cash_flow_stmt = cash_flow_dict.get(date, {})
            ratio = ratios_dict.get(date, {})
            key_metric = key_metrics_dict.get(date, {})

            revenue.append(income_stmt.get('revenue', 0))
            eps.append(income_stmt.get('eps', 0))
            fcf.append(cash_flow_stmt.get('freeCashFlow', 0))

            roe_value = ratio.get('returnOnEquity', 0)
            try:
                roe.append(float(roe_value) if roe_value not in (None, '', '0') else 0)
            except ValueError:
                logging.debug(f"Invalid ROE value on {date}: {roe_value}")
                roe.append(0)

            gross_margin.append(income_stmt.get('grossProfitRatio', 0))
            operating_margin.append(income_stmt.get('operatingIncomeRatio', 0))

            working_capital_value = income_stmt.get('totalCurrentAssets', 0) - income_stmt.get('totalCurrentLiabilities', 0)
            working_capital.append(working_capital_value)

            rd_expense.append(income_stmt.get('researchAndDevelopmentExpenses', 0))
            capex.append(abs(cash_flow_stmt.get('capitalExpenditure', 0)))

            # Use TTM PER if available, else fallback
            if per_value_ttm and per_value_ttm != 0:
                per_value = per_value_ttm
            else:
                per_value = key_metric.get('peRatio', 0)

            # Use TTM PBR if available, else fallback
            if pbr_value_ttm and pbr_value_ttm != 0:
                pbr_value = pbr_value_ttm
            else:
                pbr_value = key_metric.get('pbRatio', 0)

            per.append(float(per_value) if per_value not in (None, '', '0') else 0)
            pbr.append(float(pbr_value) if pbr_value not in (None, '', '0') else 0)

            dates.append(date)

        ttm_fcf = sum(fcf[:4]) if len(fcf) >= 4 else sum(fcf)

        return FinancialMetrics(
            revenue=revenue,
            eps=eps,
            fcf=fcf,
            ttm_fcf=ttm_fcf,
            roe=roe,
            gross_margin=gross_margin,
            operating_margin=operating_margin,
            working_capital=working_capital,
            rd_expense=rd_expense,
            capex=capex,
            per=per,
            pbr=pbr,
            dates=dates
        )
    except Exception as e:
        logging.exception(f"Error preparing financial metrics: {str(e)}")
        return None

async def get_comprehensive_financial_data(session, symbol, semaphore):
    endpoints = {
        'income_statements': f"{BASE_URL_V3}/income-statement/{symbol}?limit=20&apikey={API_KEY}",
        'cash_flow_statements': f"{BASE_URL_V3}/cash-flow-statement/{symbol}?limit=20&apikey={API_KEY}",
        'ratios': f"{BASE_URL_V3}/ratios/{symbol}?limit=20&apikey={API_KEY}",
        'ratios_ttm': f"{BASE_URL_V3}/ratios-ttm/{symbol}?apikey={API_KEY}",
        'key_metrics': f"{BASE_URL_V3}/key-metrics/{symbol}?limit=20&apikey={API_KEY}",
        'financial_growth': f"{BASE_URL_V3}/financial-growth/{symbol}?limit=20&apikey={API_KEY}",
    }

    tasks = []
    keys = []
    for key, url in endpoints.items():
        tasks.append(fetch(session, url, semaphore))
        keys.append(key)

    results_list = await asyncio.gather(*tasks)
    if any(r is None for r in results_list):
        return None
    results = dict(zip(keys, results_list))
    return results

def write_enhanced_output(filename: str, data: dict):
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"Screened {data['total_stocks']} NASDAQ stocks. "
                f"Found {data['passed_stocks_count']} qualifying stocks.\n\n")

        for stock in data['stocks']:
            f.write(f"Symbol: {stock['symbol']}\n")
            f.write(f"Company Name: {stock['company_name']}\n")
            f.write(f"Sector: {stock['sector']}\n")
            f.write(f"Market Cap: ${stock['market_cap']:,}\n")
            f.write(f"Quality Score: {stock['normalized_quality_score']:.4f}\n\n")

            f.write("Valuation Metrics:\n")
            f.write(f"  PER (Price Earnings Ratio): {stock['metrics']['per']:.2f}\n")
            f.write(f"  PBR (Price to Book Ratio): {stock['metrics']['pbr']:.2f}\n\n")

            f.write("Growth Metrics:\n")
            f.write(f"  Revenue CAGR: {stock['metrics']['revenue_cagr']:.2%}\n")
            f.write(f"  EPS CAGR: {stock['metrics']['eps_cagr']:.2%}\n")
            f.write(f"  FCF CAGR: {stock['metrics']['fcf_cagr']:.2%}\n")
            f.write(f"  Average ROE: {stock['metrics']['avg_roe']:.2%}\n")
            f.write(f"  Latest ROE: {stock['metrics']['latest_roe']:.2%}\n\n")

            f.write("Growth Quality:\n")
            f.write(f"  Revenue Consistency: {stock['growth_analysis']['revenue_consistency']:.4f}\n")
            f.write(f"  EPS Consistency: {stock['growth_analysis']['eps_consistency']:.4f}\n")
            f.write(f"  FCF Consistency: {stock['growth_analysis']['fcf_consistency']:.4f}\n")
            f.write(f"  Revenue Magnitude Score: {stock['growth_analysis']['revenue_magnitude']:.4f}\n")
            f.write(f"  EPS Magnitude Score: {stock['growth_analysis']['eps_magnitude']:.4f}\n")
            f.write(f"  FCF Magnitude Score: {stock['growth_analysis']['fcf_magnitude']:.4f}\n\n")

            f.write("Risk Assessment:\n")
            f.write(f"  Margin Risk Score: {stock['risk_assessment']['margin_risk']:.4f}\n")
            f.write(f"  Working Capital Efficiency: {stock['risk_assessment']['working_capital_efficiency']:.4f}\n\n")

            f.write("Component Scores:\n")
            for component, score in stock['component_scores'].items():
                if component in ['individual_magnitude_scores', 'individual_consistency_scores']:
                    f.write(f"  {component.replace('_', ' ').title()}:\n")
                    for metric, value in score.items():
                        f.write(f"    {metric.replace('_', ' ').title()}: {value:.4f}\n")
                elif isinstance(score, (int, float)):
                    f.write(f"  {component.replace('_', ' ').title()}: {score:.4f}\n")
                else:
                    f.write(f"  {component.replace('_', ' ').title()}: {score}\n")
            f.write("\n" + "-" * 80 + "\n\n")

def write_excel_output(filename: str, data: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Screened Stocks Results"

    # Headers
    headers = [
        "Symbol", "Company Name", "Sector", "Market Cap", "Quality Score",
        "PER", "PBR", "Revenue CAGR", "EPS CAGR", "FCF CAGR", "Avg ROE", "Latest ROE",
        "Revenue Consistency", "EPS Consistency", "FCF Consistency",
        "Revenue Magnitude", "EPS Magnitude", "FCF Magnitude",
        "Margin Risk", "WC Efficiency", "Coherence Multiplier", "Final Score (Raw)"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Populate data
    row_idx = 2
    for stock in data['stocks']:
        ws.cell(row=row_idx, column=1, value=stock['symbol'])
        ws.cell(row=row_idx, column=2, value=stock['company_name'])
        ws.cell(row=row_idx, column=3, value=stock['sector'])
        ws.cell(row=row_idx, column=4, value=stock['market_cap'])
        ws.cell(row=row_idx, column=5, value=stock['normalized_quality_score'])

        ws.cell(row=row_idx, column=6, value=stock['metrics']['per'])
        ws.cell(row=row_idx, column=7, value=stock['metrics']['pbr'])
        ws.cell(row=row_idx, column=8, value=stock['metrics']['revenue_cagr'])
        ws.cell(row=row_idx, column=9, value=stock['metrics']['eps_cagr'])
        ws.cell(row=row_idx, column=10, value=stock['metrics']['fcf_cagr'])
        ws.cell(row=row_idx, column=11, value=stock['metrics']['avg_roe'])
        ws.cell(row=row_idx, column=12, value=stock['metrics']['latest_roe'])

        ws.cell(row=row_idx, column=13, value=stock['growth_analysis']['revenue_consistency'])
        ws.cell(row=row_idx, column=14, value=stock['growth_analysis']['eps_consistency'])
        ws.cell(row=row_idx, column=15, value=stock['growth_analysis']['fcf_consistency'])
        ws.cell(row=row_idx, column=16, value=stock['growth_analysis']['revenue_magnitude'])
        ws.cell(row=row_idx, column=17, value=stock['growth_analysis']['eps_magnitude'])
        ws.cell(row=row_idx, column=18, value=stock['growth_analysis']['fcf_magnitude'])

        ws.cell(row=row_idx, column=19, value=stock['risk_assessment']['margin_risk'])
        ws.cell(row=row_idx, column=20, value=stock['risk_assessment']['working_capital_efficiency'])

        coherence_multiplier = stock['component_scores'].get('coherence_multiplier', 0.0)
        ws.cell(row=row_idx, column=21, value=coherence_multiplier)

        final_score_raw = stock['component_scores'].get('final_score', 0.0)
        ws.cell(row=row_idx, column=22, value=final_score_raw)

        row_idx += 1

    wb.save(filename)

async def screen_nasdaq(initial_filters, weights):
    start_time = time.time()
    failed_symbols = {}
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)

    normalizer = MetricNormalizer(winsorize_percentile=0.05)
    growth_analyzer = GrowthQualityAnalyzer(
        normalizer, config.get('target_rates', {'revenue': 0.20, 'eps': 0.15, 'fcf': 0.15})
    )
    risk_assessor = RiskAssessmentModule(normalizer)
    valuation_analyzer = ValuationAnalyzer(normalizer, config.get('valuation', {}))
    quality_scorer = QualityScorer(normalizer, growth_analyzer, risk_assessor, valuation_analyzer)

    connector = aiohttp.TCPConnector(limit=MAX_CONCURRENT_REQUESTS)
    async with aiohttp.ClientSession(connector=connector) as session:
        symbols_url = f"{BASE_URL_V3}/symbol/NASDAQ?apikey={API_KEY}"
        nasdaq_stocks = await fetch(session, symbols_url, semaphore)
        if not nasdaq_stocks:
            logging.error("Failed to retrieve NASDAQ symbols.")
            return

        total_stocks = len(nasdaq_stocks)
        logging.info(f"Retrieved {total_stocks} NASDAQ symbols.")

        symbols = [stock['symbol'] for stock in nasdaq_stocks]
        batches = [symbols[i:i + 100] for i in range(0, len(symbols), 100)]
        profiles = []
        for batch in batches:
            try:
                profiles_url = f"{BASE_URL_V3}/profile/{','.join(batch)}?apikey={API_KEY}"
                batch_profiles = await fetch(session, profiles_url, semaphore)
                if isinstance(batch_profiles, list):
                    profiles.extend([
                        p for p in batch_profiles
                        if isinstance(p, dict) and p.get('symbol')
                    ])
                else:
                    logging.warning(f"Invalid profile batch response: {type(batch_profiles)}")
            except Exception as e:
                logging.error(f"Error fetching profiles batch: {str(e)}")
                continue

        symbol_profile_map = {profile['symbol']: profile for profile in profiles}

        for stock in nasdaq_stocks:
            profile = symbol_profile_map.get(stock['symbol'])
            if profile and isinstance(profile.get('mktCap'), (int, float)):
                stock['sector'] = profile.get('sector', 'N/A')
                stock['companyName'] = profile.get('companyName', 'N/A')
                stock['market_cap'] = profile.get('mktCap')
            else:
                stock['sector'] = 'N/A'
                stock['companyName'] = 'N/A'
                stock['market_cap'] = 0

        filtered_stocks = [
            stock for stock in nasdaq_stocks
            if stock['market_cap'] is not None and
            isinstance(stock['market_cap'], (int, float)) and
            (initial_filters['market_cap_min'] <= stock['market_cap'] <= initial_filters['market_cap_max']) and
            (not initial_filters.get('exclude_financial_sector', False) or stock['sector'] != 'Financial Services')
        ]

        logging.info(f"Market cap screening complete. {len(filtered_stocks)} stocks passed.")

        async def analyze_stock(stock):
            symbol = stock['symbol']
            try:
                financial_data = await get_comprehensive_financial_data(session, symbol, semaphore)
                if not financial_data:
                    logging.debug(f"{symbol}: Failed to get comprehensive financial data")
                    return None

                metrics = prepare_financial_metrics(
                    financial_data['income_statements'],
                    financial_data['cash_flow_statements'],
                    financial_data['ratios'],
                    financial_data['ratios_ttm'],
                    financial_data['key_metrics'],
                    financial_data['financial_growth']
                )

                if metrics is None:
                    logging.debug(f"{symbol}: Failed to prepare financial metrics")
                    return None

                roe_criteria = initial_filters['roe']
                recent_roe_values = metrics.roe[:roe_criteria['years']]
                logging.debug(f"{symbol}: Recent ROE values: {recent_roe_values}")
                if len(recent_roe_values) < roe_criteria['years']:
                    logging.debug(f"{symbol}: Insufficient ROE history. Need {roe_criteria['years']} years")
                    return None

                avg_roe = statistics.mean(recent_roe_values)
                if avg_roe < roe_criteria['min_avg'] or any(roe < roe_criteria['min_each_year'] for roe in recent_roe_values):
                    logging.debug(f"{symbol}: Failed ROE criteria. Avg: {avg_roe:.2f}, Min required: {roe_criteria['min_avg']:.2f}")
                    return None

                quality_score, component_scores = quality_scorer.calculate_final_score(metrics, stock['market_cap'], sector=stock['sector'])

                revenue_cagr = calculate_cagr(metrics.revenue[-1], metrics.revenue[0], len(metrics.revenue))
                eps_cagr = calculate_cagr(metrics.eps[-1], metrics.eps[0], len(metrics.eps))
                fcf_cagr = calculate_cagr(metrics.fcf[-1], metrics.fcf[0], len(metrics.fcf))

                growth_analysis = {
                    'revenue_consistency': growth_analyzer.calculate_growth_consistency(metrics.revenue),
                    'eps_consistency': growth_analyzer.calculate_growth_consistency(metrics.eps),
                    'fcf_consistency': growth_analyzer.calculate_growth_consistency(metrics.fcf),
                    'revenue_magnitude': component_scores['individual_magnitude_scores']['revenue'],
                    'eps_magnitude': component_scores['individual_magnitude_scores']['eps'],
                    'fcf_magnitude': component_scores['individual_magnitude_scores']['fcf']
                }

                risk_assessment = {
                    'margin_risk': risk_assessor.calculate_margin_risk(metrics),
                    'working_capital_efficiency': risk_assessor.calculate_working_capital_efficiency(metrics)
                }

                stock_detail = {
                    'symbol': symbol,
                    'company_name': stock['companyName'],
                    'sector': stock['sector'],
                    'market_cap': stock['market_cap'],
                    'quality_score': quality_score,
                    'component_scores': component_scores,
                    'metrics': {
                        'revenue_cagr': revenue_cagr,
                        'eps_cagr': eps_cagr,
                        'fcf_cagr': fcf_cagr,
                        'avg_roe': avg_roe,
                        'latest_roe': recent_roe_values[0],
                        'per': metrics.per[0],
                        'pbr': metrics.pbr[0]
                    },
                    'growth_analysis': growth_analysis,
                    'risk_assessment': risk_assessment
                }

                return stock_detail

            except Exception as e:
                logging.exception(f"Error analyzing {symbol}: {str(e)}")
                failed_symbols[symbol] = str(e)
                return None

        semaphore_analyze = asyncio.Semaphore(config['concurrency']['max_workers'])
        async def bounded_analyze(stock):
            async with semaphore_analyze:
                return await analyze_stock(stock)

        tasks = [bounded_analyze(stock) for stock in filtered_stocks]
        results = []
        all_quality_scores = []
        for future in asyncio.as_completed(tasks):
            result = await future
            if result:
                results.append(result)
                all_quality_scores.append(result['quality_score'])

        if results:
            normalized_scores = normalizer.normalize(all_quality_scores, 'final_quality_score')
            for result, norm_score in zip(results, normalized_scores):
                result['normalized_quality_score'] = norm_score

        final_results = sorted(results, key=lambda x: x['normalized_quality_score'], reverse=True)

        timestamp = get_timestamp()
        filename_txt = f"{config['output'].get('filename_prefix', 'nasdaq_growth_stocks')}_{timestamp}.txt"
        filename_xlsx = f"{config['output'].get('filename_prefix', 'nasdaq_growth_stocks')}_{timestamp}.xlsx"

        write_enhanced_output(
            filename_txt,
            {
                'total_stocks': total_stocks,
                'passed_stocks_count': len(final_results),
                'stocks': final_results
            }
        )

        write_excel_output(
            filename_xlsx,
            {
                'total_stocks': total_stocks,
                'passed_stocks_count': len(final_results),
                'stocks': final_results
            }
        )

        end_time = time.time()
        logging.info(f"Enhanced screening complete. Runtime: {end_time - start_time:.2f} seconds")
        logging.info(f"Results written to {filename_txt} and {filename_xlsx}")

def deep_update(d, u):
    for k, v in u.items():
        if isinstance(v, dict) and k in d:
            deep_update(d[k], v)
        else:
            d[k] = v

def create_gui():
    root = tk.Tk()
    root.title("Enhanced NASDAQ Stock Screener")

    filters_frame = ttk.LabelFrame(root, text="Initial Filters")
    filters_frame.pack(fill="both", expand="yes", padx=10, pady=5)

    growth_frame = ttk.LabelFrame(root, text="Growth Quality Settings")
    growth_frame.pack(fill="both", expand="yes", padx=10, pady=5)

    quality_frame = ttk.LabelFrame(root, text="Quality Metrics")
    quality_frame.pack(fill="both", expand="yes", padx=10, pady=5)

    ttk.Label(filters_frame, text="Market Cap Min ($M):").grid(row=0, column=0, padx=5, pady=2)
    market_cap_min = ttk.Entry(filters_frame)
    market_cap_min.insert(0, str(config['initial_filters']['market_cap_min'] // 1_000_000))
    market_cap_min.grid(row=0, column=1, padx=5, pady=2)

    ttk.Label(filters_frame, text="Market Cap Max ($M):").grid(row=1, column=0, padx=5, pady=2)
    market_cap_max = ttk.Entry(filters_frame)
    market_cap_max.insert(0, str(config['initial_filters']['market_cap_max'] // 1_000_000))
    market_cap_max.grid(row=1, column=1, padx=5, pady=2)

    exclude_financial = tk.BooleanVar(value=config['initial_filters']['exclude_financial_sector'])
    ttk.Checkbutton(filters_frame, text="Exclude Financial Sector",
                    variable=exclude_financial).grid(row=2, column=0, columnspan=2, pady=2)

    ttk.Label(growth_frame, text="Min Revenue CAGR (%):").grid(row=0, column=0, padx=5, pady=2)
    rev_cagr = ttk.Entry(growth_frame)
    rev_cagr.insert(0, str(config['growth_quality']['revenue_growth']['min_cagr'] * 100))
    rev_cagr.grid(row=0, column=1, padx=5, pady=2)

    ttk.Label(growth_frame, text="Min EPS CAGR (%):").grid(row=1, column=0, padx=5, pady=2)
    eps_cagr = ttk.Entry(growth_frame)
    eps_cagr.insert(0, str(config['growth_quality']['eps_growth']['min_cagr'] * 100))
    eps_cagr.grid(row=1, column=1, padx=5, pady=2)

    def start_screening():
        try:
            updated_config = {
                'initial_filters': {
                    'market_cap_min': int(float(market_cap_min.get()) * 1_000_000),
                    'market_cap_max': int(float(market_cap_max.get()) * 1_000_000),
                    'exclude_financial_sector': exclude_financial.get()
                },
                'growth_quality': {
                    'revenue_growth': {
                        'min_cagr': float(rev_cagr.get()) / 100
                    },
                    'eps_growth': {
                        'min_cagr': float(eps_cagr.get()) / 100
                    }
                }
            }
            root.destroy()
            asyncio.run(main(updated_config))
        except ValueError as e:
            messagebox.showerror("Input Error", f"Invalid input: {str(e)}")

    ttk.Button(root, text="Start Screening", command=start_screening).pack(pady=10)
    return root

async def main(user_config=None):
    if user_config:
        deep_update(config, user_config)
    await screen_nasdaq(config['initial_filters'], config['scoring']['weights'])

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--no-gui":
        asyncio.run(main())
    else:
        root = create_gui()
        root.mainloop()
